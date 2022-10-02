import csv
import re
import string
import time
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from pydantic import BaseModel
from sectname import convert_address_to_sectcode

ROOT_DIR = Path(__file__).parent
DATA_DIR = ROOT_DIR / "moea_data"
SHEET_KEY_WORDS = ["名單", "公開資料", "工作表"]
COLUMN_KEY_WORDS = {"編號", "縣市", "地號", "使用分區"}


def has_sheet_key_word(name: str) -> bool:
    for kw in SHEET_KEY_WORDS:
        if kw in name and name != "待勘查案件名單":
            return True
    return False


class FactoryData(BaseModel):
    """
    Attributes
    ----------
    src: 來源
    no: row number in the file
    name: 工廠名稱或違章建築 (請寫名稱或貼照片)
    city: 縣市
    sectstr: 地號
    address: 地址
    status: 縣市政府會勘情形
    sect_list: 是否裝設AMI監控
    """

    src: str
    no: int
    name: Any
    city: str
    sectstr: Optional[str]
    address: Optional[str]
    status: str
    ami: Optional[str]
    sect_list: Optional[list]


class MoeaSheet:
    def __init__(self, name, sheet):
        self.name = name
        self.sheet = sheet
        self.column_names = []
        self.start_index = 0
        self._images = {}

        self._load_column_names()
        self._load_images()

    def _load_images(self):
        sheet_images = self.sheet._images
        for image in sheet_images:
            row = image.anchor._from.row + 1
            col = string.ascii_uppercase[image.anchor._from.col]
            self._images[f"{col}{row}"] = image._data

    def _load_column_names(self) -> None:
        for index, row in enumerate(self.sheet):
            cell_values = [cell.value for cell in row if cell.value]
            if len(COLUMN_KEY_WORDS.intersection(cell_values)) > 1:
                self.column_names = cell_values
                self.start_index = index + 1
                break

    def row_to_factory(self, index, row) -> FactoryData:
        cell_values = {
            column_name: cell.value for column_name, cell in zip(self.column_names, row)
        }

        cell_number = f"B{index}"
        if self._images.get(cell_number):
            name = self._images[cell_number]
        else:
            name = cell_values.get("工廠名稱或違章建築\n(請寫名稱或貼照片)") or cell_values.get("工廠名稱或違章建築")

        return FactoryData(
            src=self.name,
            no=cell_values.get("編號") or cell_values.get("No"),
            name=name,
            city=cell_values.get("市縣") or cell_values.get("縣市"),
            sectstr=cell_values.get("地號"),
            address=cell_values.get("地址") or cell_values.get("地址(去識別化)"),
            status=cell_values.get("縣市政府查處情形")
            or cell_values.get("縣市政府會勘情形")
            or cell_values.get("市縣政府查處情形"),
            ami=cell_values.get("是否裝設AMI監控") or cell_values.get("裝設AMI監控情形"),
            sect_list=None,
        )

    def get_all_factory_data(self):
        factory_data = []
        for index, row in enumerate(self.sheet):
            if index < self.start_index:
                continue
            try:
                cells = [cell.value for cell in row]
                if not any(cells):
                    continue
                parsed_row = self.row_to_factory(index, row)
                factory_data.append(parsed_row)
            except Exception as e:
                print(f"[ERROR] {e}")
                print(f"Can't parse {self.name} row: {index} data: {cells}")
                continue

        return factory_data


def get_sectcode(city: str, sectstr: Optional[str]):
    if not sectstr:
        return []
    if city not in sectstr:
        sectstr = f"{city}{sectstr}"
    full_list, simple_list = convert_address_to_sectcode(sectstr)
    result = list(map(lambda item: str(item), full_list))
    if not result:
        print(f"Failed converting {sectstr}")
        return []
    return result


def load_excel_files():
    sheets = []
    for file_path in DATA_DIR.iterdir():
        if not file_path.is_file():
            continue

        if not re.search(r"1[0-1][0|9]\d\d查處名單", file_path.stem) and not re.search(
            r"\d\d[1-9]\d\d內政部違反土地使用查處名單", file_path.stem
        ):
            continue

        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            if has_sheet_key_word(sheet_name):
                sheets.append(
                    MoeaSheet(
                        name=f"{file_path.stem}/{sheet_name}",
                        sheet=workbook[sheet_name],
                    )
                )
    return sheets

def main() -> None:
    sheets = load_excel_files()

    all_factories = []
    for sheet in sheets:
        all_factories.extend(sheet.get_all_factory_data())

    deduplicated_factories = {}
    for factory in sorted(all_factories, key=lambda x: x.src):
        if factory.sectstr:
            deduplicated_factories[factory.sectstr] = factory
        elif factory.address:
            deduplicated_factories[factory.address] = factory
        else:
            print(f"[ERROR] No factory sectstr or address in {factory}")

    for factory in deduplicated_factories.values():
        time.sleep(0.05)
        factory.sect_list = get_sectcode(factory.city, factory.sectstr)

    with open(ROOT_DIR / "moea_factories.csv", "w") as csvfile:
        writer = csv.DictWriter(
            csvfile, fieldnames=list(FactoryData.schema()["properties"].keys())
        )
        writer.writeheader()
        for factory in deduplicated_factories.values():
            writer.writerow(factory.dict())


if __name__ == "__main__":
    main()
