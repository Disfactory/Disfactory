import os
import string
from typing import List
import requests
import base64
import json
import codecs

from os import listdir
from collections import namedtuple

from openpyxl import load_workbook
from sectname import convert_address_to_sectcode, SectInfo

ROOT_DIR = os.path.dirname(__file__)
XLSX_DATA_DIR = os.path.join(ROOT_DIR, "moea_data", "xlsx")
JSON_DATA_DIR = os.path.join(ROOT_DIR, "moea_data", "json_with_sectcode")
DISFACTORY_API = "https://staging.disfactory.tw/api/sectcode"
"""
Columns:
    0. No
    1. 工廠名稱或違章建築 (請寫名稱或貼照片)
    2. 縣市
    3. 地號
    4. 地址
    5. 縣市政府會勘情形
    6. 是否裝設AMI監控
"""


class FactoryData:
    def __init__(self,
                 no,
                 name,
                 city,
                 sectstr,
                 address,
                 status,
                 ami,
                 image,
                 sect_list=None) -> None:
        self.no = no
        self.name = name
        self.city = city
        self.sectstr = sectstr
        self.address = address
        self.status = status
        self.ami = ami
        self.image = image
        self.sect_list = sect_list

    def to_dict(self):
        if self.image is not None:
            image_str = base64.b64encode(self.image).decode('utf-8')
        else:
            image_str = ""

        return {
            "no": self.no,
            "name": self.name,
            "city": self.city,
            "sectstr": self.sectstr,
            "address": self.address,
            "status": self.status,
            "ami": self.ami,
            "image_str": image_str,
            "sect_list": list(map(lambda sect_info: sect_info.to_dict(), self.sect_list))
        }


class MoeaSheet:
    def __init__(self, name, sheet):
        self.name = name
        self.sheet = sheet
        self._images = {}

        self._load_images()

    def _load_images(self):
        sheet_images = self.sheet._images
        for image in sheet_images:
            row = image.anchor._from.row + 1
            col = string.ascii_uppercase[image.anchor._from.col]
            self._images[f'{col}{row}'] = image._data()

    def get_data_list(self) -> List[FactoryData]:
        data_list = []
        for index, row in enumerate(self.sheet, start=1):
            try:
                cell_number = f"B{index}"
                if self._images.get(cell_number):
                    name = ""
                    image = self._images[cell_number]
                else:
                    name = row[1].value
                    image = None

                data_list.append(
                    FactoryData(
                        row[0].value,  # no
                        name,
                        row[2].value,  # city
                        row[3].value,  # sectstr
                        row[4].value,  # address
                        row[5].value,  # status
                        row[6].value,  # ami
                        image,  # image
                        None  # sect_list
                    ))
            except Exception as e:
                #print(f"Can't parse data {row}")
                continue

        return data_list


class MoeaXlsxFile:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(filename=file_path)
        self.sheet_names = self.wb.sheetnames
        self.unknown_sectstr = []

    def get_target_sheets(self):
        sheets = []

        for sheet_name in self.wb.sheetnames:
            if "名單" in sheet_name:
                sheets.append(MoeaSheet(sheet_name, self.wb[sheet_name]))

        return sheets

    def convert_to_factory_data_list(self) -> List[FactoryData]:
        moea_sheets = self.get_target_sheets()

        for sheet in moea_sheets:
            data_list = sheet.get_data_list()

            # add sectcode to factory data
            for item in data_list:
                item.sect_list = []
                if item.sectstr is None:
                    continue

                full_list, _ = convert_address_to_sectcode(item.sectstr)
                result = full_list

                if not result:
                    self.unknown_sectstr.append(item)
                else:
                    item.sect_list = result

        return data_list


class AddSectCodeToMoeaData:
    """
    def get_factory_id(self, sect_info: SectInfo):
        print(sect_info)
        print(f"Get the factory id by SectInfo: sect_code:{sect_info.sectcode} land_code:{sect_info.landcode}")
        result = requests.get(DISFACTORY_API, params={
            "sectcode": sect_info.sectcode,
            "landcode": sect_info.landcode
        })
        if result.status_code == 200:
            print(result.text)

    def convert_csv(self, moea_xlsx_file: MoeaXlsxFile):
        factory_data_list = moea_file.convert_to_factory_data_list()

        # Try to get the factory id
        for factory_data in factory_data_list:
            print(f"{factory_data.no}")
            for sect_info in factory_data.sect_list:
                get_factory_id(sect_info)
                pass
    """
    def convert_moea_xlsx_to_json(self, file_path):
        moea_xlsx_file = MoeaXlsxFile(file_path)
        factory_data_list = moea_xlsx_file.convert_to_factory_data_list()

        dict_data = []
        for factory_data in factory_data_list:
            dict_data.append(factory_data.to_dict())

        xlsx_file_name = os.path.basename(file_path)
        json_file_name = xlsx_file_name.replace(".xlsx", ".json")
        moea_json_file_path = os.path.join(JSON_DATA_DIR, json_file_name)
        with codecs.open(moea_json_file_path, "w", "utf-8") as fp:
            print(f"Converting {file_path} to {moea_json_file_path}")
            fp.write(json.dumps(dict_data))

    def start(self):
        file_path_list = [
            os.path.join(XLSX_DATA_DIR, f) for f in listdir(XLSX_DATA_DIR)
            if os.path.isfile(os.path.join(XLSX_DATA_DIR, f))
        ]

        for file_path in file_path_list:
            print(f"Converting xlsx file {file_path}...")
            self.convert_moea_xlsx_to_json(file_path)


if __name__ == "__main__":
    converter = AddSectCodeToMoeaData()
    converter.start()
